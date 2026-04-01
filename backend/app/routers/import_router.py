"""Import router — TestIT Excel import."""

import asyncio

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.middleware.jwt_auth import require_auth
from app.models.user import User
from app.services.testit_import_service import (
    create_job, get_job, run_import,
)

router = APIRouter(prefix="/v1/import", tags=["import"])

MAX_FILE_SIZE = 50 * 1024 * 1024  # 50 MB


@router.post("/testit")
async def import_testit(
    file: UploadFile = File(...),
    project_id: str = Query(..., description="Target project ID"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_auth),
):
    """Start TestIT Excel import as background job."""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx or .xls files are supported")

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(413, "File too large (max 50 MB)")

    job_id = create_job(project_id=project_id, filename=file.filename)
    asyncio.create_task(run_import(job_id, content, project_id))

    return {
        "job_id": job_id,
        "status": "pending",
        "message": f"Import started for file '{file.filename}'"
    }


@router.get("/jobs/{job_id}")
async def get_import_job(
    job_id: str,
    user: User = Depends(require_auth),
):
    """Get import job status and progress."""
    job = get_job(job_id)
    if not job:
        raise HTTPException(404, "Job not found")
    return job


@router.post("/testit/preview")
async def preview_testit(
    file: UploadFile = File(...),
    user: User = Depends(require_auth),
):
    """Preview first 10 rows of TestIT Excel file."""
    import openpyxl
    from io import BytesIO
    from app.services.testit_import_service import extract_title, map_priority, map_status, detect_columns

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(413, "File too large")

    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    # Read header row to detect column mapping
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(h).strip().lower() if h else "" for h in row]
        break

    col_map = detect_columns(headers)

    rows = []
    for row in ws.iter_rows(min_row=2, max_row=11, values_only=True):
        if not row or not any(row):
            continue
        padded = list(row) + [None] * max(0, len(headers) - len(row))
        title_raw = padded[col_map["title"]] if col_map["title"] is not None else None
        title, _ = extract_title(title_raw)
        rows.append({
            "external_id": padded[col_map["id"]] if col_map["id"] is not None else None,
            "folder": str(padded[col_map["folder"]])[:80] if col_map["folder"] is not None and padded[col_map["folder"]] else None,
            "title": title[:100] if title else "",
            "priority": map_priority(padded[col_map["priority"]] if col_map["priority"] is not None else None),
            "status": map_status(padded[col_map["status"]] if col_map["status"] is not None else None),
            "has_steps": bool(padded[col_map["steps"]]) if col_map["steps"] is not None else False,
        })

    return {
        "total_rows": ws.max_row - 1 if ws.max_row else 0,
        "preview": rows,
        "sheet_name": ws.title,
        "headers": headers,
        "col_map": col_map,
    }
