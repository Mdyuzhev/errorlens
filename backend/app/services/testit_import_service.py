"""TestIT Excel import service — parse and batch-create test cases."""

import logging
import re
import uuid
from datetime import datetime
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import async_session_maker
from app.models.testcase import TestCase, TestCaseFolder
from app.services.project_service import ProjectService

logger = logging.getLogger(__name__)

# ── In-memory job registry ──────────────────────────────────────
_IMPORT_JOBS: dict[str, dict] = {}


def create_job(project_id: str, filename: str) -> str:
    job_id = str(uuid.uuid4())
    _IMPORT_JOBS[job_id] = {
        "id": job_id,
        "project_id": project_id,
        "filename": filename,
        "status": "pending",
        "total": 0,
        "imported": 0,
        "skipped": 0,
        "folders_created": 0,
        "errors": [],
        "started_at": None,
        "finished_at": None,
        "progress_pct": 0,
    }
    return job_id


def get_job(job_id: str) -> dict | None:
    return _IMPORT_JOBS.get(job_id)


# ── Field mapping ───────────────────────────────────────────────

PRIORITY_MAP = {
    "высокий": "high",
    "средний": "medium",
    "низкий": "low",
    "самый низкий": "low",
}

STATUS_MAP = {
    "готов": "ready",
    "не готов": "draft",
    "требуется доработка": "needs_work",
}


# ── Column detection ─────────────────────────────────────────────

# Known header patterns for TestIT export formats
_HEADER_PATTERNS: dict[str, list[str]] = {
    "id": ["id", "ид", "№", "#", "external_id", "номер"],
    "title": ["title", "название", "name", "наименование", "тест-кейс", "тест кейс", "test case", "test_case"],
    "folder": ["location", "расположение", "раздел", "folder", "section", "путь", "path", "секция"],
    "steps": ["steps", "шаги", "step", "actions", "действия"],
    "preconditions": ["preconditions", "предусловия", "предусловие"],
    "postconditions": ["postconditions", "постусловия", "постусловие"],
    "expected": ["expected", "ожидаемый", "expected result", "ожидаемый результат"],
    "priority": ["priority", "приоритет"],
    "status": ["status", "статус", "state", "состояние"],
    "automated": ["automated", "автоматизирован", "автоматизация", "automation"],
    "tags": ["tags", "теги", "тег", "метки", "labels"],
    "author": ["author", "автор", "created by", "создал"],
    "created_at": ["created", "создан", "дата создания", "created at", "date"],
}


def detect_columns(headers: list[str]) -> dict[str, int | None]:
    """Map logical field names to column indices by matching header text."""
    result: dict[str, int | None] = {k: None for k in _HEADER_PATTERNS}

    for i, header in enumerate(headers):
        h = header.lower().strip()
        if not h:
            continue
        for field, patterns in _HEADER_PATTERNS.items():
            if result[field] is not None:
                continue
            for pattern in patterns:
                if pattern in h or h in pattern:
                    result[field] = i
                    break

    return result


def extract_title(raw: str | None) -> tuple[str, str | None]:
    if not raw:
        return "", None
    raw = str(raw).strip()
    # Match =HYPERLINK("url", "title") or HYPERLINK("url", "title")
    match = re.search(r'=?HYPERLINK\(\s*"([^"]+)"\s*,\s*"(.*)"\s*\)\s*$', raw, re.DOTALL)
    if match:
        url = match.group(1)
        title = match.group(2).replace('""', '"').strip()
        return title, url
    return raw, None


def map_priority(raw: str | None) -> str:
    if not raw:
        return "medium"
    return PRIORITY_MAP.get(str(raw).strip().lower(), "medium")


def map_status(raw: str | None) -> str:
    if not raw:
        return "draft"
    return STATUS_MAP.get(str(raw).strip().lower(), "draft")


def map_automation(raw: str | None) -> str:
    if raw and str(raw).strip().lower() in ("да", "yes", "true", "1"):
        return "Automated"
    return "Manual"


def parse_steps(steps_text: str | None, expected_text: str | None) -> list[dict]:
    if not steps_text and not expected_text:
        return []

    steps_str = str(steps_text).strip() if steps_text else ""
    expected_str = str(expected_text).strip() if expected_text else ""

    # Split by --- separator if present (multi-step in single cell)
    if "---" in steps_str or "---" in expected_str:
        action_parts = [p.strip() for p in steps_str.split("---") if p.strip()] if steps_str else []
        expected_parts = [p.strip() for p in expected_str.split("---") if p.strip()] if expected_str else []
        max_len = max(len(action_parts), len(expected_parts))
        result = []
        for i in range(max_len):
            result.append({
                "action": action_parts[i] if i < len(action_parts) else "",
                "expected": expected_parts[i] if i < len(expected_parts) else "",
                "data": None,
            })
        return result

    # Single step
    result = []
    if steps_str:
        result.append({"action": steps_str, "expected": expected_str, "data": None})
    elif expected_str:
        result.append({"action": "", "expected": expected_str, "data": None})
    return result


def parse_tags(raw: str | None) -> list[str]:
    if not raw:
        return []
    return [t.strip() for t in str(raw).split(",") if t.strip()]


def parse_folder_path(raw: str | None) -> list[str]:
    if not raw:
        return []
    return [p.strip() for p in str(raw).split("->") if p.strip()]


# ── Folder cache & creation ─────────────────────────────────────

async def get_or_create_folder(
    db: AsyncSession,
    project_id: str,
    path_parts: list[str],
    folder_cache: dict,
) -> str | None:
    if not path_parts:
        return None

    parent_id = None
    for part in path_parts:
        cache_key = (project_id, parent_id, part)
        if cache_key in folder_cache:
            parent_id = folder_cache[cache_key]
            continue

        stmt = select(TestCaseFolder).where(
            TestCaseFolder.project_id == project_id,
            TestCaseFolder.name == part,
            TestCaseFolder.parent_id == parent_id,
        )
        result = await db.execute(stmt)
        folder = result.scalar_one_or_none()

        if not folder:
            folder = TestCaseFolder(
                project_id=project_id,
                name=part,
                parent_id=parent_id,
            )
            db.add(folder)
            await db.flush()
            await db.refresh(folder)

        folder_cache[cache_key] = folder.id
        parent_id = folder.id

    return parent_id


# ── Main import ─────────────────────────────────────────────────

async def run_import(job_id: str, file_bytes: bytes, project_id: str) -> None:
    job = _IMPORT_JOBS.get(job_id)
    if not job:
        return

    job["status"] = "running"
    job["started_at"] = datetime.utcnow().isoformat()

    try:
        import openpyxl
        from io import BytesIO

        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=False)
        ws = wb.active

        total = ws.max_row - 1 if ws.max_row else 0
        job["total"] = total

        # Detect columns from header row
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(h).strip().lower() if h else "" for h in row]
            break
        col_map = detect_columns(headers)
        ncols = max(len(headers), 16)

        logger.info("Import %s: headers=%s, col_map=%s", job_id, headers, col_map)

        folder_cache: dict = {}
        imported = 0
        skipped = 0
        folders_created_set: set = set()
        errors: list[str] = []

        BATCH_SIZE = 100

        def _get(padded: list, field: str) -> Any:
            idx = col_map.get(field)
            if idx is None or idx >= len(padded):
                return None
            return padded[idx]

        async with async_session_maker() as db:
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
                if not row or not any(row):
                    continue

                try:
                    padded = list(row) + [None] * max(0, ncols - len(row))

                    ext_id = _get(padded, "id")
                    location = _get(padded, "folder")
                    title_raw = _get(padded, "title")
                    automated = _get(padded, "automated")
                    preconditions = _get(padded, "preconditions")
                    steps_raw = _get(padded, "steps")
                    postconditions = _get(padded, "postconditions")
                    expected = _get(padded, "expected")
                    priority_raw = _get(padded, "priority")
                    status_raw = _get(padded, "status")
                    author = _get(padded, "author")
                    tags_raw = _get(padded, "tags")

                    # Check duplicate
                    if ext_id:
                        dup_stmt = select(TestCase.id).where(
                            TestCase.project_id == project_id,
                            TestCase.external_id == str(ext_id),
                        )
                        dup_result = await db.execute(dup_stmt)
                        if dup_result.scalar_one_or_none():
                            skipped += 1
                            continue

                    title, external_url = extract_title(title_raw)
                    if not title:
                        errors.append(f"Row {row_idx}: empty title")
                        continue

                    path_parts = parse_folder_path(location)
                    folder_id = None
                    if path_parts:
                        folder_id = await get_or_create_folder(
                            db, project_id, path_parts, folder_cache
                        )
                        if folder_id:
                            folders_created_set.add(folder_id)

                    # Generate human_id
                    human_id = None
                    try:
                        ps = ProjectService(db)
                        human_id = await ps.next_human_id(project_id)
                    except Exception:
                        pass

                    tc = TestCase(
                        project_id=project_id,
                        human_id=human_id,
                        title=title[:500],
                        preconditions=str(preconditions).strip() if preconditions else None,
                        postconditions=str(postconditions).strip() if postconditions else None,
                        steps=parse_steps(steps_raw, expected),
                        tags=parse_tags(tags_raw),
                        priority=map_priority(priority_raw),
                        status=map_status(status_raw),
                        automation_status=map_automation(automated),
                        folder_id=folder_id,
                        external_id=str(ext_id) if ext_id else None,
                        external_url=external_url,
                        created_by=str(author).strip() if author else None,
                    )
                    db.add(tc)
                    imported += 1

                    if imported % BATCH_SIZE == 0:
                        await db.commit()
                        job["imported"] = imported
                        job["skipped"] = skipped
                        job["folders_created"] = len(folders_created_set)
                        job["progress_pct"] = round(row_idx / total * 100, 1) if total else 0

                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)[:100]}")
                    logger.warning("Import error row %d: %s", row_idx, e)

            await db.commit()

        job["imported"] = imported
        job["skipped"] = skipped
        job["folders_created"] = len(folders_created_set)
        job["errors"] = errors[:50]
        job["progress_pct"] = 100
        job["status"] = "done"
        job["finished_at"] = datetime.utcnow().isoformat()

        logger.info("Import %s done: %d imported, %d skipped, %d errors", job_id, imported, skipped, len(errors))

    except Exception as e:
        job["status"] = "error"
        job["errors"] = [str(e)]
        job["finished_at"] = datetime.utcnow().isoformat()
        logger.exception("Import %s failed", job_id)
